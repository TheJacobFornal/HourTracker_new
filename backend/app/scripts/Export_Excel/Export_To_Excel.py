from openpyxl import Workbook
from io import BytesIO
from app.scripts.Export_Data_DB import OUT_Main
from openpyxl.styles import Font
from datetime import datetime


def main(project_name):
    """Create Excel file in memory and return as BytesIO."""
    wb = Workbook()
    ws = wb.active
    ws.title = project_name

    # Set headers
    ws["A1"], ws["B1"], ws["C1"] = "Aktywność", "Osoba", "Suma godzin"

    bold_font = Font(bold=True)
    for col in ["A", "B", "C", "E"]:
        ws[f"{col}1"].font = bold_font

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["E"].width = 20

    data = OUT_Main.get_data_to_Export_Excel(project_name)

    for item in data:
        ws.append(
            [
                item["activity_name"],
                item["user_full_name"],
                item["total_combined_hours"],
            ]
        )

    ws.cell(row=1, column=5).value = "Data eksportu:"
    ws.cell(row=2, column=5).value = datetime.now().strftime("%Y-%m-%d")

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


# This block only runs if the file is executed directly
if __name__ == "__main__":
    project_name = "251501-KB10"
    print(f"Creating Excel file in memory for project: {project_name}")

    excel_file = main(project_name)

    # Optional: save to disk for testing
    test_file = f"{project_name}_{datetime.now().strftime('%Y-%m-%d')}_HourTracker.xlsx"
    with open(test_file, "wb") as f:
        f.write(excel_file.read())

    print(f"Excel file saved locally as: {test_file}")
