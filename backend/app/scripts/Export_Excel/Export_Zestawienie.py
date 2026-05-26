import datetime
import calendar
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from app.db.session import SessionLocal
from app.models.project import Project
from app.models.time_log import TimeLog


def generate(month: int, year: int, activity_ids: list[int]) -> BytesIO:
    date_from = datetime.date(year, month, 1)
    if month == 12:
        date_to_excl = datetime.date(year + 1, 1, 1)
    else:
        date_to_excl = datetime.date(year, month + 1, 1)

    with SessionLocal() as db:
        rows = (
            db.query(
                Project.name.label("project_name"),
                func.sum(TimeLog.hours).label("total_hours"),
            )
            .join(TimeLog, TimeLog.project_id == Project.id)
            .filter(
                TimeLog.activity_id.in_(activity_ids),
                TimeLog.log_date >= date_from,
                TimeLog.log_date < date_to_excl,
            )
            .group_by(Project.name)
            .order_by(func.sum(TimeLog.hours).desc())
            .all()
        )

    _MONTHS_PL = [
        "Styczeń", "Luty", "Marzec", "Kwiecień", "Maj", "Czerwiec",
        "Lipiec", "Sierpień", "Wrzesień", "Październik", "Listopad", "Grudzień",
    ]
    month_name = f"{_MONTHS_PL[month - 1]} {year}"

    wb = Workbook()
    ws = wb.active
    ws.title = "Zestawienie"

    total_all = sum(float(r.total_hours or 0) for r in rows)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1A90FF")
    center = Alignment(horizontal="center")

    headers = ["Projekt", "Suma godzin", "Udział %"]
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, row in enumerate(rows, start=2):
        hours = float(row.total_hours or 0)
        pct = (hours / total_all) if total_all else 0.0
        ws.cell(row=row_idx, column=1, value=row.project_name)
        ws.cell(row=row_idx, column=2, value=hours)
        ws.cell(row=row_idx, column=3, value=pct).number_format = "0.00%"

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14

    # metadata
    meta_row = len(rows) + 3
    ws.cell(row=meta_row, column=1, value="Okres:").font = Font(bold=True)
    ws.cell(row=meta_row, column=2, value=month_name)
    ws.cell(row=meta_row + 1, column=1, value="Data eksportu:").font = Font(bold=True)
    ws.cell(row=meta_row + 1, column=2, value=datetime.datetime.now().strftime("%Y-%m-%d"))

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream
