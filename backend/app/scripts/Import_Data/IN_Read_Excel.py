from datetime import time
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from openpyxl import load_workbook
from app.scripts.Import_Data.IN_DB import IN_db
from openpyxl.utils import column_index_from_string
import datetime
from app.scripts.Import_Data.IN_DB import IN_db_check_insert
from datetime import date


## Additional function ##
def time_to_decimal(time_obj: time) -> Decimal:
    if not isinstance(time_obj, time):
        raise ValueError("Input must be a datetime.time object")

    total_hours = (
        time_obj.hour
        + (time_obj.minute / 60)
        + (time_obj.second / 3600)
        + (time_obj.microsecond / 3600000000)
    )

    decimal_hours = Decimal(str(total_hours)).quantize(
        Decimal("0.01"), rounding=ROUND_HALF_UP
    )
    return decimal_hours


def get_start_col(year, month):
    if year < 2017 or (year == 2018 and month <= 6):  # R
        return 18
    elif year < 2020 or (year == 2020 and month <= 4):  # U
        return 21
    elif year < 2025 or (year == 2025 and month <= 5):  # V
        return 22
    elif year == 2025 and month <= 7:  # X
        return 24
    else:  # Y
        return 25


## Excel function ##
def name_surname(ws, file_name):
    Name_Surname = ws.cell(2, 1).value  # Name and Surname

    if not Name_Surname:
        Name_Surname = file_name
    if Name_Surname:
        # split on any whitespace, collapse multiple spaces
        parts = Name_Surname.strip().split()

        if len(parts) >= 2:
            surname = parts[0]
            name = parts[1]
            IN_db_check_insert.check_insert_user(name, surname)
            return name, surname

    raise ValueError(f"Invalid name format in cell A2: {Name_Surname!r}")


def get_project_map(ws, year, month):
    """Create a mapping of column numbers to project names and insert project do db if not exist"""
    start_col = get_start_col(year, month)
    col_to_project = {}

    # Project names are in row 5, every 5 columns
    for col in range(start_col, 179, 5):
        cell_value = ws.cell(5, col).value
        if cell_value:
            
            project_name = str(cell_value).split()[0]
        else:
            project_name = None

        if project_name is not None:
            project_name = str(project_name).strip()  # Convert to string
        else:
            continue

        if project_name:
            IN_db_check_insert.check_insert_project(project_name)

            # Map this column and the next 4 columns to the same project
            for i in range(5):
                if col + i <= 179:
                    col_to_project[col + i] = project_name
    return col_to_project


def get_activity(ws, col):
    activity = ws.cell(6, col).value
    return activity if activity else "Inne"


def get_date(row, month, year):
    day = row - 10
    try:
        return datetime.date(year, month, day)
    except ValueError:
        return None


## Initial Functions - checking exisitng and inserting Projects, User and Activity
def check_insert_activity(ws, year, month):
    min_col = get_start_col(year, month)
    for col in range(min_col, 200):
        activity = ws.cell(6, col).value

        if activity:
            IN_db_check_insert.check_insert_activity(activity)


## Main Functions ##
# Mode:
# Daily = Daily_logs
# Archive = Dane z archium


def one_row(ws, row, name, surname, month, year, project_map, mode):
    record_counter = 0
    for col in range(get_start_col(year, month), 179):
        time_val = ws.cell(row, col).value

        if isinstance(time_val, datetime.time) and time_val != time(0, 0):
            project = project_map.get(col)
            activity = get_activity(ws, col)
            date = get_date(row, month, year)

            if project and date:
                time_decimal = time_to_decimal(time_val)

                if mode in ("Daily", "Archive"):
                    try:
                        # print(activity, project, date, time_val, flush=True)
                        insert_func = (
                            IN_db.check_insert_daily_timelogs  # value "if" from IT below
                            if mode == "Daily"
                            else IN_db.check_insert_timeLog  # value "else" from IT above
                        )

                        insert_func(
                            name, surname, project, activity, date, time_decimal
                        )

                        record_counter += 1

                    except Exception as e:
                        print(f"❌ Error inserting time log ({mode}): {e}", flush=True)

                else:
                    missing = [
                        field
                        for field, value in {"project": project, "date": date}.items()
                        if not value
                    ]
                    if missing:
                        print(
                            f"⚠️ Skipping row for {name} {surname}: missing {', '.join(missing)}",
                            flush=True,
                        )

    return record_counter


def read_Sheet(year, month, Excel_path, mode):
    try:
        wb = load_workbook(Excel_path)
        ws = wb.active

        Excel_path = Path(Excel_path)

        file_name = Excel_path.stem

        project_map = get_project_map(ws, year, month)

        check_insert_activity(ws, year, month)
        name, surname = name_surname(ws, file_name)

        # Process each row
        record_counter = 0
        for row in range(11, 42):
            record_counter += one_row(
                ws, row, name, surname, month, year, project_map, mode
            )

        if record_counter > 0:
            print(
                "✅", name, surname, "dodano", record_counter, "rekordów!", flush=True
            )
        else:
            print(name, surname, "- Brak danych do zaimportowania!", flush=True)
        print(" ")

    except Exception as e:
        print(f"❌ Error in main: {e}")
    finally:
        wb.close()


def main_Archium(year, month, Excel_path):
    read_Sheet(
        year=year,
        month=month,
        Excel_path=Excel_path,
        mode="Archive",
    )


def main_Daily(year, month, Excel_path):
    read_Sheet(
        year=year,
        month=month,
        Excel_path=Excel_path,
        mode="Daily",
    )


if __name__ == "__main__":
    main_Daily(
        2033,
        9,
        Path(r"C:\Users\JakubFornal\Desktop\Łukomski Krzysztof.xlsm"),
    )
