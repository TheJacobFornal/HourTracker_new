# app/db/sess_remote.py
from pathlib import Path
from sqlalchemy import create_engine, text
from sqlalchemy.orm import sessionmaker
import urllib.parse as _u
from openpyxl import load_workbook

SERVER = r"10.1.69.13\EPLAN"
DATABASE = "worksheets"
USER = "ws"
PWD = "64>z*9zK@e$2H4CXX@W2\h.n&j(0b~yh"  # escape backslash for Python

_odbc = _u.quote_plus(
    f"DRIVER=ODBC Driver 17 for SQL Server;"
    f"SERVER={SERVER};DATABASE={DATABASE};UID={USER};PWD={PWD};"
    "Encrypt=no;"
)
engine = create_engine(
    f"mssql+pyodbc:///?odbc_connect={_odbc}", echo=False, future=True
)
SessionLocal = sessionmaker(bind=engine, autocommit=False, autoflush=False)


def read_Sheet(Excel_path):
    try:
        wb = load_workbook(Excel_path)
        ws = wb.active

        Excel_path = Path(Excel_path)

        for row in range(97, 100):
            correct_id = ws.cell(row, 1).value
            wrong_id = ws.cell(row, 3).value

            if correct_id and wrong_id:
                merge_projects(correct_id, wrong_id)
                print(
                    f"Processing row {row}: correct_id={correct_id}, wrong_id={wrong_id}"
                )
    except Exception as e:
        print(f"❌ Error in main: {e}")
    finally:
        wb.close()


def merge_projects(correct_id: int, wrong_id: int):
    """
    Merge wrong_id into correct_id by updating all records in time_logs and daily_logs.
    All records with project_id = wrong_id will be changed to correct_id.
    """
    sql = text(
        """
    BEGIN TRAN;

    BEGIN TRY
        -- Update time_logs
        UPDATE dbo.time_logs
        SET project_id = :correct_id
        WHERE project_id = :wrong_id;

        COMMIT TRAN;
    END TRY
    BEGIN CATCH
        ROLLBACK TRAN;
        THROW;
    END CATCH;
    """
    )

    with engine.begin() as conn:  # transaction-safe
        conn.execute(sql, {"correct_id": correct_id, "wrong_id": wrong_id})

    print(f"✅ Merged project {wrong_id} into project {correct_id} in both tables.")


if __name__ == "__main__":
    with engine.connect() as c:

        db_info = c.execute(text("SELECT DB_NAME(), SUSER_SNAME()")).all()
        print("DB info:", db_info)
        read_Sheet(
            Path(
                r"C:\Users\JakubFornal\Desktop\PROJECTS\HourTracker_new\backend\Inne\Project_id_list.xlsx"
            )
        )
